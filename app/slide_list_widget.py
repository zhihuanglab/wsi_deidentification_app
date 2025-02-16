from PySide6.QtWidgets import (QWidget, QVBoxLayout, QTableWidget, 
                              QTableWidgetItem, QHeaderView, QPushButton, QDialog, QLabel, QProgressDialog)
from PySide6.QtCore import Qt
from PySide6.QtGui import QPixmap
import tiffslide
import os
from slide_metadata_dialog import SlideMetadataDialog
from anonymize_slide import anonymize_slide
import subprocess
import hashlib
import pandas as pd
from datetime import datetime

class SlideListWidget(QTableWidget):
    # Add column names as class attributes
    COLUMN_FILENAME = "Filename"
    COLUMN_THUMBNAIL = "Thumbnail"
    COLUMN_LABEL = "Label"
    COLUMN_MACRO = "Macro"
    COLUMN_METADATA = "Metadata"  # New column
    COLUMNS = [COLUMN_FILENAME, COLUMN_THUMBNAIL, COLUMN_LABEL, COLUMN_MACRO, COLUMN_METADATA]

    def __init__(self):
        super().__init__()
        self.setup_ui()
        self.slides = []
        
    def setup_ui(self):
        # Set up the table
        self.setColumnCount(len(self.COLUMNS))
        self.setHorizontalHeaderLabels(self.COLUMNS)
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # Disable editing
        
        # Style the table
        self.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #E5E5E5;
                border: 1px solid #E5E5E5;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 5px;
                border: 1px solid #E5E5E5;
                font-weight: bold;
            }
            QPushButton {
                background-color: #0066CC;
                color: white;
                border: none;
                padding: 0px 0px;
                min-width: 100px;  /* Added min-width */
                height: 30px;      /* Increased height */
                margin: 0px;       /* Added margin */
            }
            QPushButton:hover {
                background-color: #0052A3;
            }
        """)

        header = self.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.setColumnWidth(0, 200)  # Set filename column width to 200px
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.setColumnWidth(4, 120)  # Set metadata column width to 120px
        
        # Enable sorting
        self.setSortingEnabled(True)
        
        # Connect double-click signal
        self.cellDoubleClicked.connect(self.handle_cell_double_click)
        
    def load_slides(self, folder_path):
        self.clear()
        self.setHorizontalHeaderLabels(self.COLUMNS)  # Reset column headers after clearing
        self.slides = []
        self.setRowCount(0)
        
        # Find all SVS and NDPI files
        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.svs', '.ndpi')):
                filepath = os.path.join(folder_path, filename)
                try:
                    slide = tiffslide.TiffSlide(filepath)
                    self.slides.append(slide)
                    self.add_slide_row(slide, filename)
                except Exception as e:
                    print(f"Error loading {filename}: {str(e)}")
    
    def add_slide_row(self, slide, filename):
        row = self.rowCount()
        self.insertRow(row)
        self.setRowHeight(row, 120)  # Adjusted for 100px image + padding
        
        # Add filename first
        self.setItem(row, 0, QTableWidgetItem(filename))
        
        # Add thumbnail with fixed height of 100
        thumbnail = slide.get_thumbnail((1000, 100))  # Large width to maintain aspect ratio
        pixmap = QPixmap.fromImage(thumbnail.toqimage())
        pixmap = pixmap.scaled(pixmap.width(), 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        thumbnail_item = QTableWidgetItem()
        thumbnail_item.setData(Qt.DecorationRole, pixmap)
        self.setItem(row, 1, thumbnail_item)
        
        # Add label and macro images if available
        try:
            label = slide.associated_images.get('label')
            label_item = QTableWidgetItem()
            if label:
                label_pixmap = QPixmap.fromImage(label.toqimage())
                label_pixmap = label_pixmap.scaled(300, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                label_item.setData(Qt.DecorationRole, label_pixmap)
            self.setItem(row, 2, label_item)
            
            macro = slide.associated_images.get('macro')
            macro_item = QTableWidgetItem()
            if macro:
                macro_pixmap = QPixmap.fromImage(macro.toqimage())
                macro_pixmap = macro_pixmap.scaled(300, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                macro_item.setData(Qt.DecorationRole, macro_pixmap)
            self.setItem(row, 3, macro_item)
            
            # Add metadata button with container widget for centering
            container = QWidget()
            layout = QVBoxLayout(container)
            metadata_button = QPushButton("Show Metadata")
            metadata_button.setFixedSize(100, 30)  # Set fixed size for the button
            layout.addWidget(metadata_button)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(5, 5, 5, 5)  # Added margins around the button
            metadata_button.clicked.connect(lambda checked, r=row: self.show_metadata(r))
            self.setCellWidget(row, 4, container)
            
        except Exception as e:
            print(f"Error loading associated images for {filename}: {str(e)}")
    
    def show_metadata(self, row):
        slide = self.slides[row]
        dialog = SlideMetadataDialog(slide, self)
        dialog.exec()
    
    def anonymize_all_slides(self, folder_path, options):
        # Create new folder for de-identified files
        deid_folder = folder_path + "_DEID"
        if not os.path.exists(deid_folder):
            os.makedirs(deid_folder)
        
        # Initialize lists for honest broker file
        broker_data = []
        
        # Create progress dialog
        progress = QProgressDialog("De-identifying slides...", "Cancel", 0, self.rowCount(), self)
        progress.setWindowTitle("Progress")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)  # Show immediately
        
        for row in range(self.rowCount()):
            if progress.wasCanceled():
                break
                
            filename = self.item(row, 0).text()
            progress.setLabelText(f"De-identifying: {filename}")
            progress.setValue(row)
            
            src_path = os.path.join(folder_path, filename)
            
            # Generate new filename if encryption is enabled
            if options['encrypt_filename']:
                name, ext = os.path.splitext(filename)
                hash_obj = hashlib.md5(name.encode())
                new_filename = hash_obj.hexdigest() + ext
            else:
                new_filename = filename
            
            dst_path = os.path.join(deid_folder, new_filename)
            
            try:
                # Copy file to new location
                import shutil
                shutil.copy2(src_path, dst_path)
                
                # Get label and macro images if available
                slide = self.slides[row]
                label_img = slide.associated_images.get('label')
                macro_img = slide.associated_images.get('macro')
                
                # Save images to PIL format for Excel
                label_pil = label_img.convert('RGB') if label_img else None
                macro_pil = macro_img.convert('RGB') if macro_img else None
                
                # Add to broker data
                broker_data.append({
                    'Original Filename': filename,
                    'New Filename': new_filename,
                    'Original Label Image': label_pil,
                    'Original Macro Image': macro_pil
                })
                
                # Anonymize the copied file
                anonymize_slide(dst_path)
                
            except Exception as e:
                print(f"Error anonymizing {filename}: {str(e)}")
        
        progress.setValue(self.rowCount())  # Ensure progress bar completes
        
        # Create honest broker file if enabled
        if options['create_honest_broker'] and not progress.wasCanceled():
            progress.setLabelText("Creating honest broker file...")
            
            # Create Excel workbook with images
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "secure data mapping file"
            
            # Add headers
            headers = ['Original Filename', 'New Filename', 'Original Label Image', 'Original Macro Image']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data and images
            for row_idx, data in enumerate(broker_data, 2):
                ws.cell(row=row_idx, column=1, value=data['Original Filename'])
                ws.cell(row=row_idx, column=2, value=data['New Filename'])
                
                # Add label image
                if data['Original Label Image']:
                    img_buffer = BytesIO()
                    data['Original Label Image'].save(img_buffer, format='PNG')
                    img = Image(img_buffer)
                    # Scale image to reasonable size
                    scale_factor = min(300 / img.width, 100 / img.height)
                    img.width = int(img.width * scale_factor)
                    img.height = int(img.height * scale_factor)
                    ws.add_image(img, f'C{row_idx}')
                    ws.row_dimensions[row_idx].height = max(75, img.height)
                
                # Add macro image
                if data['Original Macro Image']:
                    img_buffer = BytesIO()
                    data['Original Macro Image'].save(img_buffer, format='PNG')
                    img = Image(img_buffer)
                    # Scale image to reasonable size
                    scale_factor = min(300 / img.width, 100 / img.height)
                    img.width = int(img.width * scale_factor)
                    img.height = int(img.height * scale_factor)
                    ws.add_image(img, f'D{row_idx}')
                    ws.row_dimensions[row_idx].height = max(75, img.height)
            
            # Adjust column widths
            for col in ['A', 'B', 'C', 'D']:
                ws.column_dimensions[col].width = 40
            
            # Save workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            broker_path = os.path.join(deid_folder, "..", f"{deid_folder}_PHI_Secure_Data_Mapping_File_{timestamp}.xlsx")
            wb.save(broker_path)
    
    def handle_cell_double_click(self, row, column):
        # Only respond to metadata and macro columns
        if column == self.COLUMNS.index(self.COLUMN_METADATA):
            self.show_metadata(row)
        elif column == self.COLUMNS.index(self.COLUMN_MACRO):
            self.show_macro_image(row)
        # All other columns will do nothing
    
    def show_macro_image(self, row):
        slide = self.slides[row]
        macro = slide.associated_images.get('macro')
        if macro:
            dialog = QDialog(self)
            dialog.setWindowTitle("Macro Image")
            layout = QVBoxLayout(dialog)
            
            # Convert macro image to QPixmap
            macro_pixmap = QPixmap.fromImage(macro.toqimage())
            
            # Create label and set the image
            label = QLabel()
            label.setPixmap(macro_pixmap)
            
            layout.addWidget(label)
            dialog.exec() 